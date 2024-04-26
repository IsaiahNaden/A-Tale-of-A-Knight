import tkinter as tk
import openpyxl

root = tk.Tk()
root.title("User Profile")
root.geometry("{0}x{1}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))
root.config(bg="black")

# Add labels and entries for user input
label_name = tk.Label(root, text="What is your character's name?", fg="black")
label_name.config(font=("Times New Roman", 18, "bold"))
label_name.place(relx=0.5, rely=0.3, anchor="center")

entry_name = tk.Entry(root)
entry_name.config(font=("Times New Roman", 13, "bold"))
entry_name.place(relx=0.5, rely=0.35, anchor="center")

label_age = tk.Label(root, text="What is your age?", fg="black")
label_age.config(font=("Times New Roman", 18, "bold"))
label_age.place(relx=0.5, rely=0.45, anchor="center")

entry_age = tk.Entry(root)
entry_age.config(font=("Times New Roman", 13, "bold"))
entry_age.place(relx=0.5, rely=0.5, anchor="center")

label_gender = tk.Label(root, text="What is your gender? (1 for male, 2 for female)", fg="black")
label_gender.config(font=("Times New Roman", 18, "bold"))
label_gender.place(relx=0.5, rely=0.6, anchor="center")

entry_gender = tk.Entry(root)
entry_age.config(font=("Times New Roman", 13, "bold"))
entry_gender.place(relx=0.5, rely=0.65, anchor="center")

button_done = tk.Button(root, text="Done", command=lambda: save_user_profile(entry_name(), entry_age(), entry_gender()))
button_done.place(relx=0.5, rely=0.8, anchor="center")

root.mainloop()

if button_done.click == True:
    exit()

def save_user_profile(user_name, user_age, user_gender):
    userprofile = openpyxl.load_workbook("userprofile.xlsx")
    profilesheet = userprofile.active

    profilesheet[f'A{profilesheet.max_row + 1}'] = user_name
    profilesheet[f'B{profilesheet.max_row}'] = user_age
    profilesheet[f'C{profilesheet.max_row}']  = user_gender

    userprofile.save("userprofile.xlsx")

    #print("Data saved")
