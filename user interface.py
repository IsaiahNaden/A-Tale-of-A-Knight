git import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
import openpyxl 
import pandas as pd
import pygame

def custom_message_box(title,message):
    top = tk.Toplevel(root)
    title = ("")
    top.title(title)
    top.attributes("-fullscreen", True)

    background_image = Image.open("background2.jpg")
    background_image = background_image.resize((root.winfo_screenwidth(), root.winfo_screenheight()))
    background_photo = ImageTk.PhotoImage(background_image)
    
    canvas = tk.Canvas(top, width=root.winfo_screenwidth(), height=root.winfo_screenheight())
    canvas.pack(fill="both", expand=True)
    canvas.create_image(0, 0, anchor="nw", image=background_photo)
    canvas.image = background_photo  

    label = tk.Label(canvas, text=message, font=("Helvetica", 30), bg="sky blue")
    label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

def start_game():
    newgame_button.pack_forget()
    start_game_button.pack_forget()
    exit_button.pack_forget()

    start_game_button_clicked()


def start_game_button_clicked():
    def ask_username():
        global username
        global userid

        username = entry_username.get()
        userid = entry_userid.get()
        check_profile(username, userid)

    def check_profile(username, userid):
        # Load the Excel file containing user profiles
        try:
            user_profiles = pd.read_excel('userprofiles.xlsx')
        except FileNotFoundError:
            
            return

        # Check if the user exists in the user profiles
        user_profile = user_profiles[(user_profiles['Username'] == username) & (user_profiles['UserID'] == userid)]
        if user_profile.empty:
            
            return
        if input(user_profile) == True :
            info = f"Welcome back, {username}!"
            custom_message_box("", info)
            root.after(2000, lambda: pygame_init_game(user_profile)) 


    root = tk.Tk()
    root.title("Game Login")



    canvas = tk.Canvas(root, width=300, height=200, bg="sky blue")
    canvas.pack()

    label_username = tk.Label(canvas, text="Username:")
    canvas.create_window(100, 50, window=label_username)

    entry_username = tk.Entry(canvas)
    canvas.create_window(200, 50, window=entry_username)

    label_userid = tk.Label(canvas, text="User ID:")
    canvas.create_window(100, 100, window=label_userid)

    entry_userid = tk.Entry(canvas)
    canvas.create_window(200, 100, window=entry_userid)

    Done_button = tk.Button(canvas, text="Done", command=ask_username)
    canvas.create_window(150, 150, window=Done_button)

    root.mainloop()




def new_game():
    newgame_button.pack_forget()
    start_game_button.pack_forget()
    exit_button.pack_forget()

    label_name.pack()
    entry_name.pack()
    label_error_name.pack()
    label_age.pack()
    entry_age.pack()
    label_error_age.pack()
    label_gender.pack()
    combo_gender.pack()
    label_error_gender.pack()
    button_submit.pack()

def pygame_init_game():
    pygame.init()
    ###############
    pygame.quit()

def exit_game():
    response = messagebox.askyesno("", "Are you sure you want to exit the game?")
    if response:
        root.destroy()

def get_user_name():
    user_name = entry_name.get()
    if len(user_name) > 10:
        label_error_name.config(text="*Please enter a name less than 10 characters.", fg="red", font=("Times New Roman", 15))
        return get_user_name()
    elif user_name == "":
        label_error_name.config(text="*Please enter a name.", fg="red", font=("Times New Roman", 15))
        return get_user_name()
    else:
        label_error_name.config(text="")
        return user_name

def get_user_age():
    if entry_age.get().isdigit() == False:
        label_error_age.config(text="*Please enter a number.", fg="red", font=("Times New Roman", 15))
        return get_user_age()
    elif entry_age.get() == "":
        label_error_age.config(text="*Please enter your age.", fg="red", font=("Times New Roman", 15))
        return get_user_name()
    elif int(entry_age.get()) < 1 or int(entry_age.get()) > 100:
        label_error_age.config(text="*Please enter a valid number.", fg="red", font=("Times New Roman", 15))
        return get_user_age()
    else:
        label_error_age.config(text="")
        return int(entry_age.get())

def get_user_gender():
    if combo_gender.get() == "":
        label_error_gender.config(text="*Please select a gender.", fg="red", font=("Times New Roman", 15))
        return get_user_gender()
    else:
        return combo_gender.get()

def save_user_profile():
    user_name = get_user_name()
    user_age = get_user_age()
    user_gender = get_user_gender()

    user_profile = openpyxl.load_workbook("userprofile.xlsx")
    profile_sheet = user_profile.active

    profile_sheet[f'A{profile_sheet.max_row + 1}'] = user_name
    profile_sheet[f'B{profile_sheet.max_row}'] = user_age
    profile_sheet[f'C{profile_sheet.max_row}'] = user_gender

    user_profile.save("userprofile.xlsx")
    root.destroy()

def load_user_profile(username, userid):
    try:
        user_profile = openpyxl.load_workbook("userprofile.xlsx")
        profile_sheet = user_profile.active
        for row in range(1, profile_sheet.max_row + 1):
            if profile_sheet[f'A{row}'].value == username and profile_sheet[f'B{row}'].value == userid:
                username = profile_sheet[f'A{row}'].value
                userid = profile_sheet[f'B{row}'].value
                user_profile.close()
                return 
        return None
    except FileNotFoundError:
        return None

root = tk.Tk()
root.title("")
root.attributes('-fullscreen', True)

image_path = "background2.jpg"
image = Image.open(image_path)

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

image = image.resize((screen_width, screen_height))
image_width, image_height = image.size
aspect_ratio = image_width / image_height

if screen_width / screen_height > aspect_ratio:
    new_width = int(screen_height * aspect_ratio)
    new_height = screen_height
else:
    new_width = screen_width
    new_height = int(screen_width / aspect_ratio)

image = image.resize((new_width, new_height))
photo = ImageTk.PhotoImage(image)

canvas = tk.Canvas(root, width=screen_width, height=screen_height)
canvas.pack()

canvas.create_image(screen_width // 2, screen_height // 2, image=photo, anchor=tk.CENTER)

menu = tk.Frame(root, bg="sky blue")
menu.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

newgame_button = tk.Button(menu, text="New Game", command=new_game, bg="white", fg="black")
newgame_button.config(font=("Times New Roman", 18))
newgame_button.pack(pady=10)

start_game_button = tk.Button (menu, text="Start Game", command=start_game, bg="white", fg="black")
start_game_button.config(font=("Times New Roman", 18))
start_game_button.pack(pady=10)

exit_button = tk.Button(menu, text="Exit", command=exit_game, bg="white", fg="black")
exit_button.config(font=("Times New Roman ", 18))
exit_button.pack(pady=10)


label_name = tk.Label(menu, text="What is your character's name?", fg="black", bg="sky blue", font=("Times New Roman", 18))
entry_name = tk.Entry(menu)
label_error_name = tk.Label(menu, text="", fg="red", bg="sky blue")

label_age = tk.Label(menu, text="What is your age?", fg="black", bg="sky blue", font=("Times New Roman", 18))
entry_age = tk.Entry(menu)
label_error_age = tk.Label(menu, text="", fg="red", bg="sky blue")

label_gender = tk.Label(menu, text="What is your gender?", fg="black", bg="sky blue", font=("Times New Roman", 18))
combo_gender = ttk.Combobox(menu, values=["Male", "Female"], state="readonly")
label_error_gender = tk.Label(menu, text="", fg="red", bg="sky blue")

button_submit = tk.Button(menu, text="Done", command=save_user_profile, fg="black", bg="white", font=("Times New Roman", 18))

label_username = tk.Label(menu, text="What is your character's name?", fg="black", bg="sky blue", font=("Times New Roman", 18))
entry_username = tk.Entry(menu)
label_error_username = tk.Label(menu, text="", fg="red", bg="sky blue")

label_userid = tk.Label(menu, text="What is your ID?", fg="black", bg="sky blue", font=("Times New Roman", 18))
entry_userid = tk.Entry(menu)
label_error_userid = tk.Label(menu, text="", fg="red", bg="sky blue")

buttton_start = tk.Button(menu, text="Done", command=start_game, fg="black", bg="white", font=("Times New Roman", 18))

root.mainloop()
